"""
Stadtjeföhl Auftritte – Streamlit + SQLite
"""

from __future__ import annotations
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

APP_TITLE = "🎵 Stadtjeföhl Auftritte – Techniker Verfügbarkeiten"
DB_PATH = Path("gigs.db")

# Excel-Datei im selben Ordner wie app.py
EXCEL_PATH = Path(__file__).parent / "Auftritte Stadtjeföhl.xlsx"
SHEET_NAME = "geplant 2526"

# Farben
PRIMARY = "#ff2b95"

# -----------------------------
# DB-Layer
# -----------------------------

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def init_db():
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS availability (
                event_id TEXT NOT NULL,
                tech_name TEXT NOT NULL,
                status TEXT NOT NULL CHECK(status IN ('Kann','Kann nicht','Unsicher')),
                updated_at TEXT NOT NULL,
                UNIQUE(event_id, tech_name)
            );
            """
        )

def upsert_availability(event_id: str, tech_name: str, status: str):
    ts = datetime.utcnow().isoformat()
    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO availability(event_id, tech_name, status, updated_at)
            VALUES(?,?,?,?)
            ON CONFLICT(event_id, tech_name) DO UPDATE SET
                status=excluded.status,
                updated_at=excluded.updated_at
            """,
            (event_id, tech_name.strip(), status, ts),
        )

def load_availability() -> pd.DataFrame:
    with get_conn() as conn:
        df = pd.read_sql_query("SELECT * FROM availability", conn)
    return df

# -----------------------------
# Excel-Layer
# -----------------------------

@dataclass
class ColumnMap:
    col_date: str = "Datum"
    col_time: str = "Uhrzeit"
    col_event: str = "Event"
    col_address: str = "Adresse"
    col_venue: str = "Location"
    col_city: str = "Stadt"
    col_duration: str = "Dauer"

@st.cache_data(show_spinner=False)
def read_excel() -> pd.DataFrame:
    path_obj = Path(EXCEL_PATH)
    if not path_obj.exists():
        raise FileNotFoundError(f"Excel-Datei nicht gefunden: {path_obj}")
    df = pd.read_excel(path_obj, sheet_name=SHEET_NAME)
    df.columns = [str(c).strip() for c in df.columns]
    return df

def build_events_df(df: pd.DataFrame, cmap: ColumnMap) -> pd.DataFrame:
    out = pd.DataFrame()
    # Merke die Excel-Zeile (Header = Zeile 1)
    out["xl_row"] = (df.reset_index(drop=True).index + 2).astype(int)

    out["date"] = pd.to_datetime(df[cmap.col_date], errors="coerce").dt.date
    out["time"] = df[cmap.col_time].astype(str)
    out["event"] = df[cmap.col_event]
    out["address"] = df[cmap.col_address]
    out["venue"] = df[cmap.col_venue]
    out["city"] = df[cmap.col_city]
    out["duration"] = df[cmap.col_duration]

    out["event_id"] = (
        out["date"].astype(str)
        + "|" + out["time"]
        + "|" + out["event"]
        + "|" + out["venue"]
        + "|" + out["city"]
    ).apply(lambda x: str(abs(hash(x))))

    out["display_dt"] = out.apply(lambda r: f"{r['date']} {r['time']}", axis=1)
    out = out.sort_values(["date","time"]).reset_index(drop=True)
    return out

# -----------------------------
# Excel write-back (Julian / Valentin)
# -----------------------------

def write_status_to_excel(excel_path: Path, sheet_name: str, xl_row: int, tech_name: str, status: str) -> None:
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt nicht gefunden: {sheet_name}")
    ws = wb[sheet_name]

    col_index = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if isinstance(val, str) and val.strip().lower() == tech_name.strip().lower():
            col_index = c
            break
    if col_index is None:
        raise ValueError(f"Spalte '{tech_name}' nicht gefunden.")

    ws.cell(row=xl_row, column=col_index).value = status
    wb.save(excel_path)

# -----------------------------
# App
# -----------------------------

def main():
    st.set_page_config(page_title="Stadtjeföhl – Auftritte", page_icon="🎵", layout="wide")

    # Kleines Header-Bild oben im Content
    HEADER_IMAGE = "https://stadtjefoehl.de/gallery_gen/e89a9bad2aae7e59e5ff2d16af1d1bc4_1932x964_0x0_1932x966_crop.jpg?ts=1754470199"
    st.image(HEADER_IMAGE, width=250)
    st.markdown(f"<h1 style='color:#111; text-align:left; margin-top: 10px;'>{APP_TITLE}</h1>", unsafe_allow_html=True)

    # Styling
    st.markdown(f"""
        <style>
        .stApp {{ background: linear-gradient(180deg, #fafafb 0%, #eef0f7 100%); }}
        .block-container {{ padding-top: 0.75rem; max-width: 1100px; }}

        /* Sidebar schwarz mit Logo */
        [data-testid="stSidebar"] {{
            background-color: #000000;
            padding-top: 1rem;
        }}
        [data-testid="stSidebar"] * {{
            color: #ffffff !important;
        }}

        /* Sidebar Input-Felder */
        [data-testid="stSidebar"] input {{
            background-color: #222 !important;
            color: #fff !important;
            border: 1px solid #555 !important;
            border-radius: 6px !important;
            padding: 6px 10px !important;
        }}
        [data-testid="stSidebar"] input::placeholder {{
            color: #aaa !important;
        }}

        .streamlit-expanderHeader {{ 
            color: #111 !important; 
            font-weight: 800 !important; 
            font-size: 1.15rem !important; 
            letter-spacing: 0.3px; 
        }}
        .stButton>button {{ background: {PRIMARY}; color:#fff; border:0; border-radius:12px; padding:8px 14px; }}
        .stButton>button:hover {{ filter: brightness(1.06); }}
        .stRadio > div[role="radiogroup"] label {{
            background: #fdf1f7; border:1px solid {PRIMARY}; color:#e0006b;
            border-radius:999px; padding:6px 12px; margin-right:8px;
        }}
        .st-expander {{ background: #ffffff; border: 1px solid #e6e8f0;
                        border-radius: 14px; box-shadow: 0 2px 10px rgba(16,24,40,0.05); }}
        .stDataFrame {{ background: #fff !important; }}
        .stDataFrame thead tr th {{ color: #111 !important; }}
        .stDataFrame tbody tr td {{ color: #333 !important; }}
        </style>
    """, unsafe_allow_html=True)

    init_db()

    # Sidebar: Logo + Name
    with st.sidebar:
        st.image("logo.png", width=160)
        st.header("👤 Dein Name")
        tech_name = st.text_input("Name für Eintragung", value="", placeholder="Dein Name")

    # Events laden
    try:
        cmap = ColumnMap()
        events = build_events_df(read_excel(), cmap)
    except Exception as e:
        st.error(f"Konnte Excel nicht laden: {e}")
        return

    avail = load_availability()

    st.subheader("📅 Termine")
    for idx, row in events.iterrows():
        eid = row["event_id"]
        date_str = row["date"].strftime("%d.%m.%Y") if hasattr(row["date"], 'strftime') else str(row["date"])
        title = f"{date_str} — {row['time']} — {row['event']} — {row['venue']} ({row['city']})"

        with st.expander(title, expanded=False):
            if isinstance(row.get("address"), str) and row["address"].strip():
                st.markdown(f"**📍 Adresse:** {row['address']}")
            dur_text = str(row.get("duration") or "").strip()
            if dur_text:
                st.markdown(f"**⏱ Dauer:** {dur_text}")

            a_df = avail[avail["event_id"] == eid].copy()
            counts = a_df.groupby("status").size().reindex(["Kann","Unsicher","Kann nicht"], fill_value=0)
            st.markdown(f"""
                <div style='display:flex; gap:8px; flex-wrap:wrap; margin:.25rem 0 .5rem;'>
                  <span style='font-weight:600; padding:6px 10px; border-radius:999px; border:1px solid #a7f3d0; background:#ecfdf5; color:#065f46;'>✅ Kann: {counts['Kann']}</span>
                  <span style='font-weight:600; padding:6px 10px; border-radius:999px; border:1px solid #fde68a; background:#fffbeb; color:#92400e;'>❔ Unsicher: {counts['Unsicher']}</span>
                  <span style='font-weight:600; padding:6px 10px; border-radius:999px; border:1px solid #fecaca; background:#fef2f2; color:#991b1b;'>❌ Kann nicht: {counts['Kann nicht']}</span>
                </div>
                """, unsafe_allow_html=True)

            if not a_df.empty:
                a_df = a_df.sort_values("tech_name")[["tech_name","status","updated_at"]]
                a_df.rename(columns={"tech_name":"Name","status":"Status","updated_at":"Aktualisiert (UTC)"}, inplace=True)
                st.dataframe(a_df, use_container_width=True, hide_index=True)
            else:
                st.caption("Noch keine Einträge.")

            if tech_name:
                status = st.radio(
                    f"Dein Status für {title}",
                    options=["Kann","Unsicher","Kann nicht"],
                    horizontal=True,
                    key=f"radio-{eid}-{idx}",
                )
                if st.button("Speichern", key=f"save-{eid}-{idx}"):
                    upsert_availability(eid, tech_name, status)
                    if tech_name.strip().lower() in {"julian","valentin"}:
                        try:
                            write_status_to_excel(EXCEL_PATH, SHEET_NAME, int(row["xl_row"]), tech_name.strip(), status)
                            st.success("Gespeichert (App + Excel).")
                        except Exception as ex:
                            st.warning(f"Gespeichert in App. Excel-Update fehlgeschlagen: {ex}")
                    else:
                        st.success("Gespeichert.")
                    st.cache_data.clear()
                    st.rerun()
            else:
                st.warning("Bitte gib links in der Seitenleiste deinen Namen ein.")

if __name__ == "__main__":
    main()

# requirements.txt
# streamlit>=1.37
# pandas>=2.1
# openpyxl>=3.1
